# coding=utf8
"""
    Command lines: Enter a city of either Oslo, Bergen or Stavanger.
                   Then optionally enter weeks where you would like a vertical line to appear on the graph.
                   Commands are separated by a space.
                   If no commands are given, the graph shown will be a totall of all traffic in all of Norway

    @ Sandra Moen
"""

import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

NUMBER_OF_WEEKS_IN_YEAR, YEARS = 52, [2013, 2014, 2015, 2016, 2017]
worksheets, vertical_lines = [], []
user_input_places = "norway"
print("Accepts command lines first: [Bergen, Oslo or Stavanger], second: set of vertical lines in weeks\nNo commands shows total traffic of all of Norway")

def getWorksheetUserEnteredCommands():
    if len(sys.argv) >= 2: # TODO: remove?
        for i in range(2, len(sys.argv)):
            vertical_lines.append(int(sys.argv[i]))
    user_input_places = sys.argv[1:][0].lower()
    if not user_input_places == 'oslo' and not user_input_places == 'bergen' and not user_input_places == 'stavanger':
        print("Error: Invalid command, please choose from 'Oslo', 'Stavanger' or 'Bergen'")

    wb2 = load_workbook('Ukestrafikk 2013-2017 utvalgte punkter Bergen - Stavanger - Oslo.xlsx')
    if user_input_places == 'bergen': ws = wb2['Bergen']
    elif user_input_places == 'oslo': ws = wb2['Oslo']
    elif user_input_places == 'stavanger': ws = wb2['Stavanger']
    return ws, user_input_places

def getWorksheetsUserEnteredNoCommands():
    worksheets, workbook = [], load_workbook('Ukeverdiger utvalgte punkter hele landet 2013-2017.xlsx')
    for ws in workbook:
        worksheets.append(ws)
    return worksheets

# Input: a single year (int), ws (worksheet)
# Output: list of total number of traffic for 52 weeks of the year (int[])
def getData(year, ws): 
    weeks = [0] * NUMBER_OF_WEEKS_IN_YEAR
    for i in range(1, ws.max_row):
        if ws['A' + str(i)].value == year:        
            for j in range(NUMBER_OF_WEEKS_IN_YEAR):
                if j < 25: # xlm counts A-Z, then AA-AZ ...
                    if not isinstance(ws[chr(66+j) + str(i)].value, int): continue # chr(66) = B in ascii
                    weeks[j] += int(ws[chr(66+j) + str(i)].value)
                elif j < 51: # xlm counts AA-AZ, then BA-BZ ...
                    if not isinstance(ws[chr(65) + chr(65+j-25) + str(i)].value, int): continue # chr(66) = B in ascii
                    weeks[j] += int(ws[chr(65) + chr(65+j-25) + str(i)].value)
                else: # xlm counts BA-BZ, then CA-CZ ...
                    if not isinstance(ws[chr(66) + chr(65) + str(i)].value, int): continue # chr(66) = B in ascii
                    weeks[j] += int(ws[chr(66) + chr(65) + str(i)].value)
    return weeks

def drawGraph(years, data): # Input: List of years (String[]). Each year contain a total amount of traffic data per week
    ticks, x = [], []
    for i in range(0, 52):
        ticks.append("Week " + str(i+1))
        x.append(i)
    x = np.array(x)

    colors=["#FF0000", "#800000", "#FFFF00", "#808000", "#00FF00", "#008000", "#00FFFF", "#008080", "#0000FF", "#000080", "#FF00FF", "#800080", "#f58231", "#aaffc3"]
    for i in range(0, len(years)):
        y = np.array(data[i])
        label = years[i]
        plt.plot(x, y, label=label, color=colors[i])

    plt.xticks(x, ticks)
    plt.xticks(rotation=45)
    plt.title('Weekly traffic from the NPRA 2013-2017 in ' + user_input_places.title())
    plt.ylabel("Traffic per week")
    plt.legend(loc='upper left')
    plt.grid(axis='y', linestyle='-')
    plt.grid(axis='x', linestyle='-')
    for xc in vertical_lines:
        plt.axvline(x=xc, color='magenta', linestyle='--')
    plt.show()

def main():
    if len(sys.argv)>1: # if console commands were given
        total_years, (worksheet, user_input_places) = [], getWorksheetUserEnteredCommands()
        for year in YEARS:
            total_years.append(getData(year, worksheet))
    else:
        total_years = [[0]*NUMBER_OF_WEEKS_IN_YEAR for i in range(len(YEARS))] # initialize empty list of lists
        worksheets = getWorksheetsUserEnteredNoCommands()
        for worksheet in worksheets:
            data_years = [[0]*NUMBER_OF_WEEKS_IN_YEAR for i in range(len(YEARS))] # initialize empty list of lists
            for year in YEARS:
                data_years[year - YEARS[0]] = getData(year, worksheet)
                for week in range(NUMBER_OF_WEEKS_IN_YEAR):
                    total_years[year - YEARS[0]][week] += data_years[year - YEARS[0]][week]
    drawGraph(YEARS, total_years)

main()
