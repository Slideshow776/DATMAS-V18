from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

wb2 = load_workbook('C:/Users/Slideshow/Dropbox/School/Masteroppgave/Git/DATMAS-V18/Trafikkregistrering Statens Vegvesen/Aars og Maanedsdøgn trafikk  2002-2015.xlsx')
ws = wb2['Trafikkdata']

def getData(year):    
    months = [] # january = 0, february = 1, ...
    for i in range(0, 12):
        months.append(0)

    for i in range(2, ws.max_row):
        if ws['D' + str(i)].value == year:        
            for j in range(0, len(months)):
                if not isinstance(ws[chr(69+j) + str(i)].value, int): # chr(69) = E in ascii
                    continue
                months[j] += int(ws[chr(69+j) + str(i)].value)
    return months

def drawGraph(data):
    ticks = ["Januar", "Februar", "Mars", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Desember"]
    x = np.array([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    colors=["#FF0000", "#800000", "#FFFF00", "#808000", "#00FF00", "#008000", "#00FFFF", "#008080", "#0000FF", "#000080", "#FF00FF", "#800080", "#f58231", "#aaffc3"]
    for i in range(2, len(data)+2):
        y = np.array(data[i-2])
        if i < 10:
            label = "200" + str(i)
        else:
            label = "20" + str(i)
        plt.plot(x, y, label=label, color=colors[i-2])

    plt.xticks(x, ticks)
    plt.title('Månedsdøgntrafikk fra Statens Vegvesen 2002-2015')
    plt.ylabel("Gjennomsnittlig trafikk per døgn")
    plt.legend(loc='upper left')
    plt.grid(axis='y', linestyle='-')
    plt.grid(axis='x', linestyle='-')
    plt.show()

data_years = []
for i in range(2002, 2016): # data tilgjengelig år 2002-2015
    data_years.append(getData(i))
drawGraph(data_years)
