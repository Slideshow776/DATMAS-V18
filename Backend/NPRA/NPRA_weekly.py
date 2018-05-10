# coding=utf8
"""
    Command lines: (1) Enter a city of either Oslo, Bergen or Stavanger.
                   (2) Enter a year of either 2013, 2014, 2015, 2016 or 2017
                   Commands are separated by a space.
                   If no commands are given, the graph shown will be a totall of all traffic in all of Norway

    @author: Sandra Moen    
"""

import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np


class NPRA_weekly:
    def __init__(self, filepath, legend=True):
        self.filepath = filepath
        self.NUMBER_OF_WEEKS_IN_YEAR, self.years = 52, [2013, 2014, 2015, 2016, 2017]
        self.worksheets = []
        self.user_input_places = "norway"

        if len(sys.argv)>1: # if console commands were given
            total_years, (worksheet, self.user_input_places) = [], self.getWorksheetUserEnteredCommands()
            for year in self.years:
                total_years.append(self.getData(year, worksheet))
        else:
            total_years = [[0]*self.NUMBER_OF_WEEKS_IN_YEAR for i in range(len(self.years))] # initialize empty list of lists
            worksheets = self.getWorksheetsUserEnteredNoCommands()
            for worksheet in worksheets:
                data_years = [[0]*self.NUMBER_OF_WEEKS_IN_YEAR for i in range(len(self.years))] # initialize empty list of lists
                for year in self.years:
                    data_years[year - self.years[0]] = self.getData(year, worksheet)
                    for week in range(self.NUMBER_OF_WEEKS_IN_YEAR):
                        total_years[year - self.years[0]][week] += data_years[year - self.years[0]][week]
        
        self._FIGURE = self.drawGraph(self.years, total_years, legend)

    def getWorksheetUserEnteredCommands(self):
        if len(sys.argv) > 1:
            self.user_input_places = sys.argv[1:][0].lower()
            if not self.user_input_places == 'oslo' and not self.user_input_places == 'bergen' and not self.user_input_places == 'stavanger':
                print("Error: Invalid command, please choose from 'Oslo', 'Stavanger' or 'Bergen'")
            if len(sys.argv) > 2:
                command = int(sys.argv[1:][1])
                if command not in self.years: 
                    print('Error: invalid year entered: Accepts [2013, 2014, 2015, 2016, 2017]')
                    sys.exit()
                self.years = [command]                

        wb2 = load_workbook(self.filepath)
        if self.user_input_places == 'bergen': ws = wb2['Bergen']
        elif self.user_input_places == 'oslo': ws = wb2['Oslo']
        elif self.user_input_places == 'stavanger': ws = wb2['Stavanger']
        return ws, self.user_input_places

    def getWorksheetsUserEnteredNoCommands(self):
        worksheets, workbook = [], load_workbook(self.filepath)
        for ws in workbook:
            worksheets.append(ws)
        return worksheets

    # Input: a single year (int), ws (worksheet)
    # Output: list of total number of traffic for 52 weeks of the year (int[])
    def getData(self, year, ws, week_start=0, week_end=None,): 
        if not week_end: week_end = self.NUMBER_OF_WEEKS_IN_YEAR
        weeks = [0] * self.NUMBER_OF_WEEKS_IN_YEAR
        for i in range(1, ws.max_row):
            if ws['A' + str(i)].value == year:        
                for j in range(week_start, week_end):
                    if j < 25: # xlm counts A-Z, then AA-AZ ...
                        if not isinstance(ws[chr(66+j) + str(i)].value, int): continue # chr(66) = B in ascii
                        weeks[j] += int(ws[chr(66+j) + str(i)].value)
                    elif j < 51: # xlm counts AA-AZ, then BA-BZ ...
                        if not isinstance(ws[chr(65) + chr(65+j-25) + str(i)].value, int): continue # chr(66) = B in ascii
                        weeks[j] += int(ws[chr(65) + chr(65+j-25) + str(i)].value)
                    else: # xlm counts BA-BZ, then CA-CZ ...
                        if not isinstance(ws[chr(66) + chr(65) + str(i)].value, int): continue # chr(66) = B in ascii
                        weeks[j] += int(ws[chr(66) + chr(65) + str(i)].value)
        return weeks

    def drawGraph(self, years, data, legend=True): # Input: List of years (String[]). Each year contain a total amount of traffic data per week
        if self.user_input_places == 'norway': plt.figure(7)
        elif self.user_input_places == 'oslo': plt.figure(8)
        elif self.user_input_places == 'bergen': plt.figure(9)
        elif self.user_input_places == 'stavanger': plt.figure(10)
        ticks, x = [], []
        for i in range(0, 52):
            ticks.append("Week " + str(i+1))
            x.append(i)
        x = np.array(x)

        colors=["#FF0000", "#800000", "#FFFF00", "#808000", "#00FF00", "#008000", "#00FFFF", "#008080", "#0000FF", "#000080", "#FF00FF", "#800080", "#f58231", "#aaffc3"]
        for i in range(0, len(years)):
            y = np.array(data[i])
            label = str(years[i])
            if not legend: label = '_' + str(label)
            plt.plot(x, y, label=label, color=colors[i])

        plt.xticks(x, ticks)
        plt.xticks(rotation=45)
        plt.title('Weekly traffic from the NPRA 2013-2017 in ' + self.user_input_places.title())
        plt.ylabel("Traffic per week")
        plt.legend(loc='upper left')
        plt.grid(axis='y', linestyle='-')
        plt.grid(axis='x', linestyle='-')
        if self.user_input_places == 'norway': figure = plt.figure(7)
        elif self.user_input_places == 'oslo': figure = plt.figure(8)
        elif self.user_input_places == 'bergen': figure = plt.figure(9)
        elif self.user_input_places == 'stavanger': figure = plt.figure(10)
        figure.patch.set_facecolor('#fff7ff')
        return figure

    def get_specific_graph_(self, user_input_places, legend=True): # accepts: 'norway', 'bergen', 'stavanger' and 'oslo'
        self.user_input_places = user_input_places
        
        if self.user_input_places == 'norway':
            total_years = [[0]*self.NUMBER_OF_WEEKS_IN_YEAR for i in range(len(self.years))] # initialize empty list of lists
            worksheets = self.getWorksheetsUserEnteredNoCommands()
            for worksheet in worksheets:
                data_years = [[0]*self.NUMBER_OF_WEEKS_IN_YEAR for i in range(len(self.years))] # initialize empty list of lists
                for year in self.years:
                    data_years[year - self.years[0]] = self.getData(year, worksheet)
                    for week in range(self.NUMBER_OF_WEEKS_IN_YEAR):
                        total_years[year - self.years[0]][week] += data_years[year - self.years[0]][week]
        else: # if console commands were given
            total_years, (worksheet, self.user_input_places) = [], self.getWorksheetUserEnteredCommands()
            for year in self.years:
                total_years.append(self.getData(year, worksheet))

        return self.drawGraph(self.years, total_years, legend)

    def get_graph(self):
        return self._FIGURE

def main():
    print("Accepts command lines: first: [Bergen, Oslo or Stavanger], second [2013, 2014, 2015, 2016, 2017]\nNo commands shows total traffic of all of Norway of all of available years")
    NPRA_weekly(
        './Ukestrafikk 2013-2017 utvalgte punkter Bergen - Stavanger - Oslo.xlsx'
        ).get_graph()
    plt.show()

if __name__ == '__main__':
    main()
