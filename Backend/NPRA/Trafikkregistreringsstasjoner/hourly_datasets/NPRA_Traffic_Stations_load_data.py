# coding=utf8

import sys, csv, time
from datetime import datetime
from openpyxl import load_workbook
import matplotlib.pyplot as plt, numpy as np

HOURLY_FILE_NAME = 'Stavanger/1100009 HILLEVÅGTUNNELEN 2013-2017.xlsx'
COORDINATES_STAVANGER_FILE_NAME = 'Stavanger/times nivå 1 STAVANGER.csv'
COORDINATES_BERGEN_FILE_NAME = 'Bergen/times nivå 1 BERGEN.csv'
COORDINATES_OSLO_FILE_NAME = 'Oslo/times nivå 1 OSLO.csv'
stavanger_stations, bergen_stations, oslo_stations = [], [], []

def get_variance_of_data(year, filename, field, hour_from, hour_to, # weeksdays range from 0-6, 0 is monday
         weekday_from, weekday_to, month_from, month_to):

    time_start = time.time()
    workbook = load_workbook(filename)
    years = []
    for sheet_year in workbook.sheetnames:
        years.append(query_data(sheet_year, filename, field, hour_from, hour_to, # weeksdays range from 0-6, 0 is monday
         weekday_from, weekday_to, month_from, month_to))
    
    # TODO: test her om i samme ukenummer, hvis så summer sammen ukedagene
    
    # --------------------------------------------------------------------

    test = []
    for y in years:
        dsum = 0
        for d in y:
            dsum += d.get_vehicles()
        test.append(dsum)
    var = np.var(test, ddof=1)
    print(var)
    print('time: ', time.time() - time_start)
    return var

    
    """print('test: ', workbook.worksheets[0])
    print('test2: ', workbook.worksheets[len(workbook.worksheets)-1])
    print('test3: ', workbook.sheetnames)
    print('years in file: ', len(workbook.worksheets))"""

    #return np.var(np.array(vehicles), ddof=1)

def query_data(year, filename, field, hour_from, hour_to, # weeksdays range from 0-6, 0 is monday
         weekday_from, weekday_to, month_from, month_to):
    query_results = []
    for d in get_data_hourly(year, filename):
        if (
            d.get_field() == field and
            hour_from <= d.get_date().hour <= hour_to and
            weekday_from <= d.get_date().weekday() <= weekday_to and
            month_from <= d.get_date().month <= month_to
        ) : query_results.append(d)
    return query_results

def get_data_hourly(year, filename): # returns data in a list of 'Traffic_recording_station' objects
    try:
        workbook = load_workbook(filename)
        worksheet = workbook[str(year)]
        stations = []
        
        startid = int(HOURLY_FILE_NAME.find('/') + 1)
        endid = int(HOURLY_FILE_NAME.find('20'))
        id = HOURLY_FILE_NAME[startid:endid]
        
        for i in range(7, worksheet.max_row):
            hour = int(worksheet['B' + str(i)].value[:2])
            if hour == 24: hour = 0
            stations.append(
                Traffic_recording_station(
                    id,
                    datetime( # date
                        int(worksheet['A' + str(i)].value[6:]),  # year
                        int(worksheet['A' + str(i)].value[3:5]), # month
                        int(worksheet['A' + str(i)].value[:2]),  # day
                        hour
                    ),
                    int(worksheet['C' + str(i)].value), # field
                    int(worksheet['D' + str(i)].value), # vehicles
                    None,
                    None
                )
            )
        return stations
    except: print('Error: Something went wrong loading the data from file or year')

def get_all_coordinates(stavanger, bergen, oslo):
    coordinates = []
    coordinates.append(_get_coordinates(stavanger))
    coordinates.append(_get_coordinates(bergen))
    coordinates.append(_get_coordinates(oslo))

    pcoordinates = []
    for c in coordinates:
        for d in c:
            pcoordinates.append(d[1:])

    new_coords = []
    for p in pcoordinates:
        new_coords.append([float(p[0]), float(p[1])])

    return new_coords

def _print_all(year, filename):
    for d in get_data_hourly(str(year), filename):
        print(d.get_all())
    print("\nYou ran the 'print_all' method, it prints all the data read from", HOURLY_FILE_NAME)

def _combine_hourly_and_coordinates(year, filename):
    data = get_data_hourly(str(year), filename)
    coordinates = _get_coordinates(COORDINATES_STAVANGER_FILE_NAME)
    for d in data:
        for c in coordinates:
            if c[0] == d.get_id():
                d.set_longitude(c[1])
                d.set_latitude(c[2])
                break
        print(d.get_all())

def _get_coordinates(filename):
    try:
        with open(filename, 'r') as f:
            reader = csv.reader(f, delimiter=';')
            next(reader, None)  # skip the headers
            coordinates = list(reader)
        return coordinates
    except:
        print('Error loading coordinates from ', filename)

class Traffic_recording_station:
    def __init__(self, id, date, field, vehicles, longitude, latitude):
        self.id = id                #               Int
        self.date = date            #               Datetime
        self.field = field          # 1 V 2,        Int
        self.vehicles = vehicles    # >= 0,         Int
        self.longitude = longitude  #               Float
        self.latitude = latitude    #               Float

    def set_id(self, id): self.id = id
    def set_date(self, date): self.date = date
    def set_field(self, field): self.field = field
    def set_vehicles(self, vehicles): self.vehicles = vehicles
    def set_longitude(self, longitude): self.longitude = longitude
    def set_latitude(self, latitude): self.latitude = latitude

    def get_id(self): return self.id
    def get_date(self): return self.date
    def get_field(self): return self.field
    def get_vehicles(self): return self.vehicles
    def get_longitude(self): return self.longitude
    def get_latitude(self): return self.latitude
    def get_all(self):
        return (
            [
                self.id,
                self.date,
                self.field,
                self.vehicles,
                self.longitude,
                self.latitude
            ]            
        )

def main():
    #data = get_data_hourly(2017, HOURLY_FILE_NAME)
    print(get_variance_of_data(
        2017, HOURLY_FILE_NAME, 1,  # year, filename, field
        8, 8,                       # hour_from, hour_to
        0, 0,                       # weekday_from, weekday_to
        0, 3)                       # month_from, month_to
    )

if __name__ == '__main__':
    main()
