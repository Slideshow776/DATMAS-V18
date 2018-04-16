# coding=utf8

import sys, datetime
from openpyxl import load_workbook
import matplotlib.pyplot as plt, numpy as np

FILE_NAME = 'Koordinater.xlsx'

def main():
    for d in get_data():
        print(d.get_all())
    print("\nYou ran the main method, it prints all the data read from", FILE_NAME)

def get_data(): # returns data in as a list of 'Traffic_recording_station' objects
    workbook = load_workbook(FILE_NAME)
    worksheet = workbook['Ark2']

    stations = []

    for i in range(4, worksheet.max_row):
        stations.append(
            Traffic_recording_station(
                worksheet['A' + str(i)].value,
                worksheet['B' + str(i)].value,
                worksheet['C' + str(i)].value,
                worksheet['D' + str(i)].value,
                worksheet['E' + str(i)].value,
                worksheet['F' + str(i)].value
            )
        )
    return stations

class Traffic_recording_station:
    def __init__(self, number, name, referance, north, east, height):
        self.number = number
        self.name = name
        self.referance = referance
        self.north = north
        self.east = east
        self.height = height

    def get_number(self): return self.number
    def get_name(self): return self.name
    def get_referance(self): return self.referance
    def get_north(self): return self.north
    def get_east(self): return self.east
    def get_height(self): return self.height
    def get_all(self):
        return (
            str(self.number) + ", " +
            self.name + ", " +
            self.referance + ", " +
            str(self.north) + ", " +
            str(self.east) + ", " +
            str(self.height)
        )

if __name__ == '__main__':
    main()
