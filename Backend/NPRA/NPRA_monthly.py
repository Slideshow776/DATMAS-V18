# coding=utf8

"""
    WIP: Command line arguments to get the wanted year.
         Pass year data to drawGraph in order to get the correct label on the graph.

    @ Sandra Moen
"""

import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np


class NPRA_monthly:
    def __init__(self):
        self.user_input = "norway"
        wb2 = load_workbook('Aars og Maanedsdøgn trafikk  2002-2015.xlsx')
        ws = wb2['Trafikkdata'] 
        data_years = []
        for i in range(2002, 2016): # data tilgjengelig år 2002-2015
            data_years.append(self.getData(i, ws, self.getPlacesBasedOnUserCommand()))
        self._FIGURE = self.drawGraph(data_years)

    def getPlacesBasedOnUserCommand(self):
        places = []
        if len(sys.argv[1:]) > 0: 
            self.user_input = sys.argv[1:][0].lower()
            if not self.user_input == 'oslo' and not self.user_input == 'bergen':
                print("Error: Invalid command, please choose 'Oslo' or 'Bergen'")
                sys.exit()
            bergen_places = ['TRENGEREIDTUNNELEN', 'SONGSTAD', 'TUNNELEN', 'TAKVAM', 'INDRE ARNA FV276', 'BJØRKHAUGTUNNELEN', 'BLINDHEIMTUNNELEN', 'GAMLE NYGÅRDSBRO', 'CHRISTIESGT', 'HEIANE BRU', 'VALLAHEIENE', 'NESTTUN TUNNEL', 'NESTTUNVEIEN SØR', 'FJØSANGER V/BOMST.(K', 'NYGÅRD', 'NYGÅRD, KRINGSJÅV.', 'FLØYFJELLTUNNELEN', 'FLØYFJ NORDGÅENDE', 'FLØYFJ SØRGÅENDE', 'AMALIE/MUNKEBOTTENT', 'SANDVIKSVEIEN', 'EIDSVÅGTUNNELEN', 'VÅGSBOTN FV 267', 'VÅGSBOTN', 'EIKÅSTUNNELEN', 'EIKÅS', 'KNARVIK', 'DAMSGÅRDSVEIEN BOM', 'DAMSGÅRDTUNNELEN', 'LYDERHORNSVEI', 'HARAFJELLTUNNELEN', 'SOTRABRUA VEST', 'KOLLTVEITTUNNELEN', 'TORBORG NEDREAASGT.', 'TROLDHAUGTUNNEL' ,'LAGUNEN' ,'HÅVARDSTUN' ,'GULLBOTN (KLIMA)' ,'ISDALSTØ FARTSTAVLE', 'FANA V/KIRKEVOLL SK', 'KJØKKELVIKVEIEN', 'LODDEFJORD NORD', 'ØVREGATEN', 'BØNESSKOGEN', 'BØNES', 'LØVSTAKKTUNNELEN', 'BJØRGEVN.SANDEIDET', 'KNAPPETUNNELEN SANDEIDET', 'SANDEIDET', 'KRÅKENES', 'STRAUME BRO', 'BJØRGEVEIEN STRAUME', 'STRAUMEVEIEN', 'YTREBYGDSVN. VED DOLVIK SØR', 'YTREBYGDSVN. VED DOLVIK NORD', 'KNAPPETUN.DOLVIK', 'KNAPPETUN.SANDEIDET', 'KNAPPETUN.STRAUME', 'RAVNANGER', 'SALHUS FARTSTAVLE', 'OSTERØYBRUA', 'HAUKELAND', 'MIDTTUNTUNNEL', 'SKJOLDSKIFTET NORD', 'SKJOLD', 'STORETVEITVEIEN BOM', 'BJØRNSONSGATEN', 'MICHAEL KROHNGT.BOM', 'MANNSVERK', 'KALFARBAKKEN', 'TORGET', 'BRYGGEN']
            oslo_places = ['E6 V/KARIHAUGEN', 'SVARTDALSTUNNELEN', 'EV 18 V/ MASTEMYR', 'E18-MOSSEV V/FISKEV', 'FESTNINGTUNNEL', 'MARITIM-510B', 'AMMERUD', 'ST.RINGV V/NYDALSBR', 'VATERLANDTUNNELEN']
            if self.user_input == 'bergen': places = bergen_places
            elif self.user_input == 'oslo': places = oslo_places
        return places

    def getData(self, year, ws, places):    
        months = [] # january = 0, february = 1, ...
        for i in range(0, 12):
            months.append(0)  
        
        if not places: # do all of the places available
            for i in range(2, ws.max_row):
                places.append(ws['B' + str(i)].value)

        for i in range(2, ws.max_row):
            if ws['B' + str(i)].value in places:
                if ws['D' + str(i)].value == year:        
                    for j in range(0, len(months)):
                        if not isinstance(ws[chr(69+j) + str(i)].value, int): continue # chr(69) = E in ascii
                        months[j] += int(ws[chr(69+j) + str(i)].value)
        return months

    def drawGraph(self, data):
        ticks = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "Desember"]
        x = np.array([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

        colors=["#FF0000", "#800000", "#FFFF00", "#808000", "#00FF00", "#008000", "#00FFFF", "#008080", "#0000FF", "#000080", "#FF00FF", "#800080", "#f58231", "#aaffc3"]
        for i in range(2, len(data)+2):
            y = np.array(data[i-2])
            if i < 10: label = "200" + str(i)
            else: label = "20" + str(i)
            plt.plot(x, y, label=label, color=colors[i-2])

        plt.xticks(x, ticks)
        plt.title('Average daily traffic by month from the NPRA 2002-2015 in ' + self.user_input.title())
        plt.ylabel("Average traffic by day")
        plt.legend(loc='upper left')
        plt.grid(axis='y', linestyle='-')
        plt.grid(axis='x', linestyle='-')
        return plt.figure(1)

    def get_specific_graph_(self, user_input): # accepts: 'norway', 'bergen' and 'oslo'
        self.user_input = user_input
        wb2 = load_workbook('Aars og Maanedsdøgn trafikk  2002-2015.xlsx')
        ws = wb2['Trafikkdata'] 
        data_years = []
        for i in range(2002, 2016): # data tilgjengelig år 2002-2015
            data_years.append(self.getData(i, ws, self.getPlacesBasedOnUserCommand()))
        return self.drawGraph(data_years)
        
    def get_graph(self):
        return self._FIGURE

def main():
    print("Accepts command lines, first: [Bergen OR Oslo]")
    NPRA_monthly().get_graph()
    plt.show()

if __name__ == '__main__':
    main()
