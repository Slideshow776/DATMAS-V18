# coding=utf8

import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import datetime


class Ruter:
    def __init__(self, filepath):
        self.filepath = filepath
        dates, passengers = self._getData()
        self._FIGURE = self._drawGraph(dates, passengers)

    def _isDate(self, date):
        try:
            datetime.datetime.strptime(date, '%d.%m.%Y')
            return True
        except ValueError:
            return False

    def _getData(self, ):
        workbook = load_workbook(self.filepath)
        worksheet = workbook['Antall påstigende per dag_Oslo']

        years_dates, years_passengers = [], []
        dates, passengers = [], []
        for i in range(5, worksheet.max_row+1):
            if self._isDate(worksheet['A' + str(i)].value):
                dates.append(worksheet['A' + str(i)].value)
                passengers.append(worksheet['B' + str(i)].value)
            else:
                years_dates.append(dates)
                years_passengers.append(passengers)
                dates, passengers = [], []
        return years_dates, years_passengers

    def _drawGraph(self, dates, passengers):
        plt.figure(2)
        x_365 = []
        for i in range(365):
            x_365.append(i)
        x_365 = np.array(x_365)
        x_366 = []
        for i in range(366):
            x_366.append(i)
        x_366 = np.array(x_366)
        x_58 = []
        for i in range(58):
            x_58.append(i)
        x_58 = np.array(x_58)
        x_ticks = dates[0]

        y_2015 = passengers[0]
        y_2016 = passengers[1]
        y_2017 = passengers[2]
        y_2018 = passengers[3]

        plt.xticks(x_365, x_ticks)
        plt.xticks(rotation=45)
        plt.plot(x_365, y_2015, label="2015")
        plt.plot(x_366, y_2016, label="2016")
        plt.plot(x_365, y_2017, label="2017")
        plt.plot(x_58, y_2018, label="2018")

        plt.title('Passengers traveling with Ruter in Oslo county')
        plt.ylabel("Passengers")
        plt.legend(loc='upper left')
        plt.grid(axis='y', linestyle='-')
        plt.grid(axis='x', linestyle='-')
        figure = plt.figure(2)
        figure.patch.set_facecolor('#fff7ff')
        return figure

    def get_graph(self):
        return self._FIGURE

def main():
    Ruter('./Antall påstigende per dag_Oslo_2015_2017.xlsx')
    plt.show()

if __name__ == '__main__':
    main()