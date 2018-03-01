import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

def get_data():
    wb2 = load_workbook(r'C:\Users\Slideshow\Dropbox\School\DATMAS-V18\Fra Lars\ILI_tall_2016_17.xlsx') # TODO: change this to relative path?
    ws = wb2['ILI_tall_2016_17']

    x, y = [], []
    for i in range(2, ws.max_row+1):
        x.append(i) # in order to get the correct x axis...
        y.append(ws['B' + str(i)].value * 100)
    return x, y

def draw_graph(x, y):
    ticks = [] # in order to get the correct x axis...
    for i in range(40, 53):
        ticks.append(str(i))
    for i in range(1, 23):
        ticks.append(str(i))

    plt.xticks(x, ticks)
    plt.plot(np.array(x), np.array(y))

    plt.title('Percent of FLS that was diagnosed with influenza 2016-2017')
    plt.xlabel("week")
    plt.ylabel("Percent")
    plt.grid(axis='y', linestyle='-')
    plt.grid(axis='x', linestyle='-')
    plt.show()

def main():    
    x, y = get_data()
    draw_graph(x, y)
    